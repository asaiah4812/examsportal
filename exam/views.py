from django.shortcuts import render, redirect, reverse
from . import forms, models
from django.db.models import Sum
from django.contrib.auth.models import Group
from django.http import HttpResponseRedirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.conf import settings
from datetime import date, timedelta
from django.db.models import Count, Q
from django.core.mail import send_mail
from teacher import models as TMODEL
from student import models as SMODEL
from teacher import forms as TFORM
from student import forms as SFORM
from django.http import HttpResponse
from django.contrib import messages
from django.urls import reverse
import io
import openpyxl
import csv
from .forms import DepartmentForm, SemesterForm
import pandas as pd
from django.contrib.auth import get_user_model

User = get_user_model()

def home_view(request):
    if request.user.is_authenticated:
        return HttpResponseRedirect('afterlogin')  
    return render(request, 'exam/index.html')

def is_teacher(user):
    return user.groups.filter(name='TEACHER').exists()

def is_student(user):
    return user.groups.filter(name='STUDENT').exists()

def afterlogin_view(request):
    if is_student(request.user):      
        return redirect('student/student-dashboard')
    elif is_teacher(request.user):
        accountapproval = TMODEL.Teacher.objects.all().filter(user_id=request.user.id, status=True)
        if accountapproval:
            return redirect('teacher/teacher-dashboard')
        else:
            return render(request, 'teacher/teacher_wait_for_approval.html')
    else:
        return redirect('admin-dashboard')

def adminclick_view(request):
    if request.user.is_authenticated:
        return HttpResponseRedirect('afterlogin')
    return HttpResponseRedirect('adminlogin')

@login_required(login_url='adminlogin')
def admin_dashboard_view(request):
    dict = {
        'total_student': SMODEL.Student.objects.all().count(),
        'total_teacher': TMODEL.Teacher.objects.all().filter(status=True).count(),
        'total_course': models.Course.objects.all().count(),
        'total_question': models.Question.objects.all().count(),
    }
    return render(request, 'exam/admin_dashboard.html', context=dict)

@login_required(login_url='adminlogin')
def admin_department_view(request):
    departments = models.Department.objects.annotate(
        course_count=Count('course'),
        student_count=Count('student')
    ).order_by('name')
    
    form = DepartmentForm()
    
    if request.method == 'POST':
        form = DepartmentForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Program added successfully!')
            return redirect('admin-department')
        else:
            messages.error(request, 'Please correct the errors below.')
    
    # Handle search parameter
    search_query = request.GET.get('search', '')
    if search_query:
        departments = departments.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query)
        )
    
    context = {
        'departments': departments,
        'form': form,
    }
    return render(request, 'exam/admin_department.html', context=context)

@login_required(login_url='adminlogin')
def delete_department_view(request, pk):
    department = models.Department.objects.get(id=pk)
    department.delete()
    return redirect('admin-department')

@login_required(login_url='adminlogin')
def update_department_view(request, pk):
    department = models.Department.objects.get(id=pk)
    form = DepartmentForm(instance=department)
    if request.method == 'POST':
        form = DepartmentForm(request.POST, instance=department)
        if form.is_valid():
            form.save()
            return redirect('admin-department')
        else:
            messages.error(request, 'Department not updated')
            return redirect('admin-department')
    context = {
        'form': form,
    }
    return render(request, 'exam/admin_department.html', context=context)

@login_required(login_url='adminlogin')
def admin_semester_view(request):
    semesters = models.Semester.objects.all()
    form = SemesterForm()
    if request.method == 'POST':
        form = SemesterForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Semester added successfully')
            return redirect('admin-semester')
        else:
            messages.error(request, 'Semester not added')
            return redirect('admin-semester')
    context = {
        'semesters': semesters,
        'form': form,
    }
    return render(request, 'exam/admin_semester.html', context=context)

@login_required(login_url='adminlogin')
def delete_semester_view(request, pk):
    semester = models.Semester.objects.get(id=pk)
    semester.delete()
    return redirect('admin-semester')

@login_required(login_url='adminlogin')
def admin_teacher_view(request):
    dict = {
        'total_teacher': TMODEL.Teacher.objects.all().filter(status=True).count(),
        'pending_teacher': TMODEL.Teacher.objects.all().filter(status=False).count(),
        'salary': TMODEL.Teacher.objects.all().filter(status=True).aggregate(Sum('salary'))['salary__sum'],
    }
    return render(request, 'exam/admin_teacher.html', context=dict)

@login_required(login_url='adminlogin')
def admin_view_teacher_view(request):
    teachers = TMODEL.Teacher.objects.all().filter(status=True)
    return render(request, 'exam/admin_view_teacher.html', {'teachers': teachers})

@login_required(login_url='adminlogin')
def update_teacher_view(request, pk):
    teacher = TMODEL.Teacher.objects.get(id=pk)
    user = TMODEL.User.objects.get(id=teacher.user_id)
    userForm = TFORM.TeacherUserForm(instance=user)
    teacherForm = TFORM.TeacherForm(request.FILES, instance=teacher)
    mydict = {'userForm': userForm, 'teacherForm': teacherForm}
    if request.method == 'POST':
        userForm = TFORM.TeacherUserForm(request.POST, instance=user)
        teacherForm = TFORM.TeacherForm(request.POST, request.FILES, instance=teacher)
        if userForm.is_valid() and teacherForm.is_valid():
            user = userForm.save()
            user.set_password(user.password)
            user.save()
            teacherForm.save()
            return redirect('admin-view-teacher')
    return render(request, 'exam/update_teacher.html', context=mydict)

@login_required(login_url='adminlogin')
def delete_teacher_view(request, pk):
    teacher = TMODEL.Teacher.objects.get(id=pk)
    user = User.objects.get(id=teacher.user_id)
    user.delete()
    teacher.delete()
    return HttpResponseRedirect('/admin-view-teacher')

@login_required(login_url='adminlogin')
def admin_view_pending_teacher_view(request):
    teachers = TMODEL.Teacher.objects.all().filter(status=False)
    return render(request, 'exam/admin_view_pending_teacher.html', {'teachers': teachers})

@login_required(login_url='adminlogin')
def approve_teacher_view(request, pk):
    teacherSalary = forms.TeacherSalaryForm()
    if request.method == 'POST':
        teacherSalary = forms.TeacherSalaryForm(request.POST)
        if teacherSalary.is_valid():
            teacher = TMODEL.Teacher.objects.get(id=pk)
            teacher.salary = teacherSalary.cleaned_data['salary']
            teacher.status = True
            teacher.save()
        else:
            print("form is invalid")
        return HttpResponseRedirect('/admin-view-pending-teacher')
    return render(request, 'exam/salary_form.html', {'teacherSalary': teacherSalary})

@login_required(login_url='adminlogin')
def reject_teacher_view(request, pk):
    teacher = TMODEL.Teacher.objects.get(id=pk)
    user = User.objects.get(id=teacher.user_id)
    user.delete()
    teacher.delete()
    return HttpResponseRedirect('/admin-view-pending-teacher')

@login_required(login_url='adminlogin')
def admin_view_teacher_salary_view(request):
    teachers = TMODEL.Teacher.objects.all().filter(status=True)
    return render(request, 'exam/admin_view_teacher_salary.html', {'teachers': teachers})

@login_required(login_url='adminlogin')
def admin_student_view(request):
    dict = {
        'total_student': SMODEL.Student.objects.all().count(),
    }
    return render(request, 'exam/admin_student.html', context=dict)

@login_required(login_url='adminlogin')
def admin_view_student_view(request):
    students = SMODEL.Student.objects.all().select_related('user', 'department')
    departments = models.Department.objects.all()  # Add this for the department filter
    
    # Handle search and filter parameters
    search_query = request.GET.get('search', '')
    department_id = request.GET.get('department', '')
    
    if search_query:
        students = students.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(matric_no__icontains=search_query)
        )
    
    if department_id:
        students = students.filter(department_id=department_id)
    
    context = {
        'students': students,
        'departments': departments,  # Pass departments to template
    }
    return render(request, 'exam/admin_view_student.html', context)

@login_required(login_url='adminlogin')
def update_student_view(request, pk):
    student = SMODEL.Student.objects.get(id=pk)
    user = SMODEL.User.objects.get(id=student.user_id)
    userForm = SFORM.StudentUserForm(instance=user)
    studentForm = SFORM.StudentForm(request.FILES, instance=student)
    mydict = {'userForm': userForm, 'studentForm': studentForm}
    if request.method == 'POST':
        userForm = SFORM.StudentUserForm(request.POST, instance=user)
        studentForm = SFORM.StudentForm(request.POST, request.FILES, instance=student)
        if userForm.is_valid() and studentForm.is_valid():
            user = userForm.save()
            user.set_password(user.password)
            user.save()
            studentForm.save()
            return redirect('admin-view-student')
    return render(request, 'exam/update_student.html', context=mydict)

@login_required(login_url='adminlogin')
def delete_student_view(request, pk):
    student = SMODEL.Student.objects.get(id=pk)
    user = User.objects.get(id=student.user_id)
    user.delete()
    student.delete()
    return HttpResponseRedirect('/admin-view-student')

@login_required(login_url='adminlogin')
def admin_bulk_upload_students_view(request):
    """
    Admin can upload a CSV or XLSX file to create multiple student accounts.
    The file should have columns: first_name, last_name, username, email, matric_no, department, address, mobile
    The default password will be the student's registration number (matric_no).
    """
    departments = models.Department.objects.all()
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        created = 0
        errors = []
        
        try:
            # Check file extension to determine how to read it
            file_extension = file.name.split('.')[-1].lower()
            
            if file_extension == 'csv':
                # Read CSV file
                decoded_file = file.read().decode('utf-8').splitlines()
                reader = csv.DictReader(decoded_file)
                rows = list(reader)
            elif file_extension in ['xlsx', 'xls']:
                # Read Excel file
                df = pd.read_excel(file)
                rows = df.to_dict('records')
            else:
                errors.append("Unsupported file format. Please upload a CSV or XLSX file.")
                context = {
                    'departments': departments,
                    'errors': errors
                }
                return render(request, 'exam/bulk_upload_students.html', context)
            
            # Process rows
            for idx, row in enumerate(rows, start=2):
                try:
                    # Convert all values to strings before stripping
                    first_name = str(row.get('first_name') or '').strip()
                    last_name = str(row.get('last_name') or '').strip()
                    email = str(row.get('email') or '').strip()
                    username = str(row.get('username') or '').strip()
                    department_name = str(row.get('department') or '').strip()
                    address = str(row.get('address') or '').strip()
                    mobile = str(row.get('mobile') or '').strip()

                    if not (first_name and last_name and username and email and username and department_name):
                        errors.append(f"Row {idx}: Missing required fields.")
                        continue

                    # Check for existing user
                    if User.objects.filter(username=username).exists():
                        errors.append(f"Row {idx}: Username '{username}' already exists.")
                        continue

                    # Get department
                    try:
                        department = models.Department.objects.get(name__iexact=department_name)
                    except models.Department.DoesNotExist:
                        errors.append(f"Row {idx}: Department '{department_name}' not found.")
                        continue

                    # Create user
                    user = User.objects.create_user(
                        username=username,
                        password=username,  # Default password is registration number
                        first_name=first_name,
                        last_name=last_name,
                        email=email
                    )
                    # Add to STUDENT group
                    group, _ = Group.objects.get_or_create(name='STUDENT')
                    user.groups.add(group)

                    # Create student profile
                    student = SMODEL.Student.objects.create(
                        user=user,
                        department=department,
                        address=address,
                        mobile=mobile
                    )
                    created += 1
                except Exception as e:
                    errors.append(f"Row {idx}: {str(e)}")
                    
        except Exception as e:
            errors.append(f"Error reading file: {str(e)}")

        context = {
            'departments': departments,
            'success': f"Successfully imported {created} students." if created else None,
            'errors': errors if errors else None
        }
        return render(request, 'exam/bulk_upload_students.html', context)

    return render(request, 'exam/bulk_upload_students.html', {
        'departments': departments
    })


@login_required(login_url='adminlogin')
def download_sample_students_csv(request):
    """
    Download a sample XLSX file for student bulk upload.
    """
    # Create sample data
    data = {
        'first_name': ['John', 'Jane'],
        'last_name': ['Doe', 'Smith'],
        'username': ['johndoe', 'janesmith'],
        'email': ['john@example.com', 'jane@example.com'],
        'username': ['johndoe', 'janesmith'],
        'department': ['Computer Science', 'Mathematics'],
        'address': ['123 Main St', '456 Elm St'],
        'mobile': ['08012345678', '08087654321']
    }
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Create in-memory Excel file
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Students', index=False)
    
    # Prepare response
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="sample_students.xlsx"'
    
    return response

@login_required(login_url='adminlogin')
def admin_course_view(request):
    return render(request, 'exam/admin_course.html')

@login_required(login_url='adminlogin')
def admin_add_course_view(request):
    departments = models.Department.objects.all()
    semesters = models.Semester.objects.all()
    courseForm = forms.CourseForm()
    if request.method == 'POST':
        courseForm = forms.CourseForm(request.POST)
        if courseForm.is_valid():
            courseForm.save()
        else:
            print("form is invalid")
        return redirect('/admin-view-course')
    return render(request, 'exam/admin_add_course.html', {'courseForm': courseForm, 'departments': departments, 'semesters': semesters})

@login_required(login_url='adminlogin')
def admin_activate_course(request, pk):
    course = models.Course.objects.get(id=pk)
    course.active = True
    course.save()
    messages.success(request, 'Course activated successfully')
    return redirect('/admin-view-course')

@login_required(login_url='adminlogin')
def admin_deactivate_course(request, pk):
    courses = models.Course.objects.all()
    course = models.Course.objects.get(id=pk)
    course.active = False
    course.save()
    messages.success(request, 'Course deactivated successfully')
    return redirect('/admin-view-course')

@login_required(login_url='adminlogin')
def admin_view_course_view(request):
    courses = models.Course.objects.all().select_related('department', 'semester')
    departments = models.Department.objects.all()
    semesters = models.Semester.objects.all()
    
    # Handle search and filter parameters
    search_query = request.GET.get('search', '')
    department_id = request.GET.get('department', '')
    semester_id = request.GET.get('semester', '')
    status_filter = request.GET.get('status', '')
    min_year = request.GET.get('min_year', '')
    max_year = request.GET.get('max_year', '')
    min_questions = request.GET.get('min_questions', '')
    max_questions = request.GET.get('max_questions', '')
    
    if search_query:
        courses = courses.filter(course_name__icontains=search_query)
    
    if department_id:
        courses = courses.filter(department_id=department_id)
    
    if semester_id:
        courses = courses.filter(semester_id=semester_id)
    
    if status_filter == 'active':
        courses = courses.filter(active=True)
    elif status_filter == 'inactive':
        courses = courses.filter(active=False)
    
    if min_year:
        courses = courses.filter(year__gte=min_year)
    
    if max_year:
        courses = courses.filter(year__lte=max_year)
    
    if min_questions:
        courses = courses.filter(question_number__gte=min_questions)
    
    if max_questions:
        courses = courses.filter(question_number__lte=max_questions)
    
    context = {
        'courses': courses,
        'departments': departments,
        'semesters': semesters,
    }
    return render(request, 'exam/admin_view_course.html', context)

@login_required(login_url='adminlogin')
def delete_course_view(request, pk):
    course = models.Course.objects.get(id=pk)
    course.delete()
    return redirect('/admin-view-course')

@login_required(login_url='adminlogin')
def admin_question_view(request):
    courses = models.Course.objects.all()
    return render(request, 'exam/admin_question.html', {'courses': courses})

@login_required(login_url='adminlogin')
def admin_add_question_view(request):
    questionForm = forms.QuestionForm()
    courses = models.Course.objects.all()
    if request.method == 'POST':
        questionForm = forms.QuestionForm(request.POST)
        if questionForm.is_valid():
            question = questionForm.save(commit=False)
            course = models.Course.objects.get(id=request.POST.get('courseID'))
            question.course = course
            question.save()
        else:
            print("form is invalid")
        return redirect('/admin-view-question')
    return render(request, 'exam/admin_add_question.html', {'questionForm': questionForm, 'courses': courses})

@login_required(login_url='adminlogin')
def admin_bulk_upload_questions_view(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        course_id = request.POST.get('course_id')
        try:
            course = models.Course.objects.get(id=course_id)
        except models.Course.DoesNotExist:
            return render(request, 'exam/bulk_upload.html', {
                'courses': models.Course.objects.all(),
                'error': 'Selected course not found.'
            })

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            created = 0
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Expected columns: question, option1, option2, option3, option4, answer, marks
                question_text, option1, option2, option3, option4, answer, marks = row or (None,) * 7
                if not question_text:
                    continue
                if answer not in ['Option1', 'Option2', 'Option3', 'Option4']:
                    continue
                try:
                    marks_int = int(marks) if marks is not None else 1
                except Exception:
                    marks_int = 1
                models.Question.objects.create(
                    course=course,
                    question=str(question_text)[:600],
                    option1=str(option1 or '')[:200],
                    option2=str(option2 or '')[:200],
                    option3=str(option3 or '')[:200],
                    option4=str(option4 or '')[:200],
                    answer=answer,
                    marks=marks_int
                )
                created += 1

            # Optionally update course metadata
            course.question_number = models.Question.objects.filter(course=course).count()
            course.total_marks = models.Question.objects.filter(course=course).aggregate(Sum('marks'))['marks__sum'] or 0
            course.save()

            return render(request, 'exam/bulk_upload.html', {
                'courses': models.Course.objects.all(),
                'success': f'Successfully imported {created} questions into {course.course_name}.'
            })
        except Exception as e:
            return render(request, 'exam/bulk_upload.html', {
                'courses': models.Course.objects.all(),
                'error': 'Failed to process file. Ensure it follows the sample format.'
            })

    return render(request, 'exam/bulk_upload.html', {
        'courses': models.Course.objects.all()
    })

@login_required(login_url='adminlogin')
def download_sample_questions_excel(request):
    # Build an in-memory Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Questions'
    ws.append(['question', 'option1', 'option2', 'option3', 'option4', 'answer', 'marks'])
    ws.append(['What is 2 + 2?', '1', '2', '3', '4', 'Option4', 2])
    ws.append(['Capital of France?', 'Berlin', 'Madrid', 'Paris', 'Lisbon', 'Option3', 5])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    response = HttpResponse(stream.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sample_questions.xlsx"'
    return response

@login_required(login_url='adminlogin')
def admin_view_question_view(request):
    courses = models.Course.objects.all().select_related('department', 'semester')
    departments = models.Department.objects.all()
    semesters = models.Semester.objects.all()
    
    # Handle search and filter parameters
    search_query = request.GET.get('search', '')
    department_id = request.GET.get('department', '')
    semester_id = request.GET.get('semester', '')
    status_filter = request.GET.get('status', '')
    
    if search_query:
        courses = courses.filter(course_name__icontains=search_query)
    
    if department_id:
        courses = courses.filter(department_id=department_id)
    
    if semester_id:
        courses = courses.filter(semester_id=semester_id)
    
    if status_filter == 'active':
        courses = courses.filter(active=True)
    elif status_filter == 'inactive':
        courses = courses.filter(active=False)
    
    context = {
        'courses': courses,
        'departments': departments,
        'semesters': semesters,
    }
    return render(request, 'exam/admin_view_question.html', context)

@login_required(login_url='adminlogin')
def view_question_view(request, pk):
    questions = models.Question.objects.all().filter(course_id=pk)
    departments = models.Department.objects.all()
    semesters = models.Semester.objects.all()
    return render(request, 'exam/view_question.html', {'questions': questions, 'departments': departments, 'semesters': semesters})

@login_required(login_url='adminlogin')
def delete_question_view(request, pk):
    question = models.Question.objects.get(id=pk)
    question.delete()
    return redirect('/admin-view-question')

@login_required(login_url='adminlogin')
def admin_view_student_marks_view(request):
    students = SMODEL.Student.objects.all()
    departments = models.Department.objects.all()
    semesters = models.Semester.objects.all()
    return render(request, 'exam/admin_view_student_marks.html', {'students': students, 'departments': departments, 'semesters': semesters})

@login_required(login_url='adminlogin')
def admin_view_marks_view(request, pk):
    courses = models.Course.objects.all()
    departments = models.Department.objects.all()
    semesters = models.Semester.objects.all()
    response = render(request, 'exam/admin_view_marks.html', {'courses': courses, 'departments': departments, 'semesters': semesters})
    response.set_cookie('student_id', str(pk))
    return response

@login_required(login_url='adminlogin')
def admin_check_marks_view(request, pk):
    try:
        course = models.Course.objects.get(id=pk)
        student_id = request.COOKIES.get('student_id')
        departments = models.Department.objects.all()
        if not student_id:
            messages.error(request, 'No student selected. Please select a student first.')
            return redirect(reverse('admin-view-student-marks'))

        try:
            student = SMODEL.Student.objects.get(id=student_id)
        except SMODEL.Student.DoesNotExist:
            messages.error(request, 'Selected student not found.')
            return redirect(reverse('admin-view-student-marks'))
        results = models.Result.objects.filter(exam=course, student=student).order_by('-date')
        departments = models.Department.objects.all()
        semesters = models.Semester.objects.all()
        # Debug info
        print(f"Course: {course.course_name}")
        print(f"Student: {student.get_name}")
        print(f"Results count: {results.count()}")

        context = {
            'results': results,
            'course': course,
            'student': student,
            'departments': departments,
            'semesters': semesters
        }
        return render(request, 'exam/admin_check_marks.html', context)

    except models.Course.DoesNotExist:
        messages.error(request, 'Course not found.')
        return redirect(reverse('admin-view-student-marks'))
    except Exception as e:
        messages.error(request, f'An error occurred: {str(e)}')
        return redirect(reverse('admin-view-student-marks'))

def aboutus_view(request):
    departments = models.Department.objects.all()
    semesters = models.Semester.objects.all()
    return render(request, 'exam/aboutus.html', {'departments': departments, 'semesters': semesters})

def contactus_view(request):
    sub = forms.ContactusForm()
    departments = models.Department.objects.all()
    semesters = models.Semester.objects.all()
    if request.method == 'POST':
        sub = forms.ContactusForm(request.POST)
        if sub.is_valid():
            email = sub.cleaned_data['Email']
            name = sub.cleaned_data['Name']
            message = sub.cleaned_data['Message']
            send_mail(str(name) + ' || ' + str(email), message, settings.EMAIL_HOST_USER, settings.EMAIL_RECEIVING_USER, fail_silently=False)
            return render(request, 'exam/contactussuccess.html', {'departments': departments, 'semesters': semesters})
    return render(request, 'exam/contactus.html', {'form': sub, 'departments': departments, 'semesters': semesters})

# Public: Check Results from landing page
def check_results_view(request):
    courses = models.Course.objects.all().order_by('course_name')
    results = None
    searched_user = None
    selected_course = None
    departments = models.Department.objects.all()
    semesters = models.Semester.objects.all()
    if request.method == 'POST':
        username = (request.POST.get('username') or '').strip()
        course_id = request.POST.get('course_id')

        if not username:
            messages.error(request, 'Please enter your username or matric number.')
        else:
            try:
                # Accept either username or student matric number
                try:
                    user = User.objects.get(username=username)
                except User.DoesNotExist:
                    # Try via matric number
                    student_by_matric = SMODEL.Student.objects.select_related('user').get(username=username)
                    user = student_by_matric.user
                searched_user = user
                student = SMODEL.Student.objects.get(user=user)

                qs = models.Result.objects.filter(student=student).order_by('-date')
                if course_id:
                    try:
                        selected_course = models.Course.objects.get(id=course_id)
                        qs = qs.filter(exam=selected_course)
                    except models.Course.DoesNotExist:
                        messages.error(request, 'Selected course was not found.')

                results = qs
                if not qs.exists():
                    messages.info(request, 'No results found for the provided details.')
            except User.DoesNotExist:
                messages.error(request, 'No user found with that username or matric number.')
            except SMODEL.Student.DoesNotExist:
                messages.error(request, 'Student profile not found for that user.')
            except SMODEL.Student.MultipleObjectsReturned:
                messages.error(request, 'Multiple student records found for that matric number.')

    context = {
        'courses': courses,
        'results': results,
        'searched_user': searched_user,
        'selected_course': selected_course,
        'departments': departments,
        'semesters': semesters
    }
    return render(request, 'exam/check_results.html', context)

