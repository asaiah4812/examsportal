from django.shortcuts import render, redirect, reverse
from . import forms, models
from django.db.models import Sum
from django.contrib.auth.models import Group
from django.http import HttpResponseRedirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.conf import settings
from datetime import date, timedelta
from exam import models as QMODEL
from student import models as SMODEL
from exam import forms as QFORM
from django.http import HttpResponse
import openpyxl
import io
from django.db.models import Sum

# for showing signup/login button for teacher
def teacherclick_view(request):
    if request.user.is_authenticated:
        return HttpResponseRedirect('afterlogin')
    return render(request, 'teacher/teacherclick.html')

def teacher_signup_view(request):
    # Add department field to the signup form
    userForm = forms.TeacherUserForm()
    teacherForm = forms.TeacherForm()
    mydict = {'userForm': userForm, 'teacherForm': teacherForm}
    if request.method == 'POST':
        userForm = forms.TeacherUserForm(request.POST)
        teacherForm = forms.TeacherForm(request.POST, request.FILES)
        if userForm.is_valid() and teacherForm.is_valid():
            user = userForm.save()
            user.set_password(user.password)
            user.save()
            teacher = teacherForm.save(commit=False)
            teacher.user = user
            # department is now part of the TeacherForm, so it will be saved
            teacher.save()
            my_teacher_group = Group.objects.get_or_create(name='TEACHER')
            my_teacher_group[0].user_set.add(user)
            return HttpResponseRedirect('teacherlogin')
    return render(request, 'teacher/teachersignup.html', context=mydict)

def is_teacher(user):
    return user.groups.filter(name='TEACHER').exists()

@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def teacher_dashboard_view(request):
    # Only show courses and questions from the teacher's department
    teacher = models.Teacher.objects.get(user=request.user)
    department = teacher.department
    total_course = QMODEL.Course.objects.filter(department=department).count()
    total_question = QMODEL.Question.objects.filter(course__department=department).count()
    total_student = SMODEL.Student.objects.filter(department=department).count()
    context = {
        'total_course': total_course,
        'total_question': total_question,
        'total_student': total_student
    }
    return render(request, 'teacher/teacher_dashboard.html', context=context)

@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def teacher_exam_view(request):
    return render(request, 'teacher/teacher_exam.html')

@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def teacher_add_exam_view(request):
    teacher = models.Teacher.objects.get(user=request.user)
    department = teacher.department
    # Limit department choices in the form if needed
    courseForm = QFORM.CourseForm()
    # Optionally, you can set the department field initial value or limit choices
    if request.method == 'POST':
        courseForm = QFORM.CourseForm(request.POST)
        if courseForm.is_valid():
            course = courseForm.save(commit=False)
            course.department = department  # Force course to teacher's department
            course.save()
        else:
            print("form is invalid")
        return HttpResponseRedirect('/teacher/teacher-view-exam')
    return render(request, 'teacher/teacher_add_exam.html', {'courseForm': courseForm})

@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def teacher_view_exam_view(request):
    teacher = models.Teacher.objects.get(user=request.user)
    department = teacher.department
    courses = QMODEL.Course.objects.filter(department=department)
    return render(request, 'teacher/teacher_view_exam.html', {'courses': courses})

@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def delete_exam_view(request, pk):
    teacher = models.Teacher.objects.get(user=request.user)
    department = teacher.department
    try:
        course = QMODEL.Course.objects.get(id=pk, department=department)
        course.delete()
    except QMODEL.Course.DoesNotExist:
        pass
    return HttpResponseRedirect('/teacher/teacher-view-exam')

@login_required(login_url='adminlogin')
def teacher_question_view(request):
    return render(request, 'teacher/teacher_question.html')

@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def teacher_add_question_view(request):
    teacher = models.Teacher.objects.get(user=request.user)
    department = teacher.department
    questionForm = QFORM.QuestionForm()
    # Limit course choices to teacher's department
    questionForm.fields['course'].queryset = QMODEL.Course.objects.filter(department=department)
    if request.method == 'POST':
        questionForm = QFORM.QuestionForm(request.POST)
        questionForm.fields['course'].queryset = QMODEL.Course.objects.filter(department=department)
        if questionForm.is_valid():
            question = questionForm.save(commit=False)
            # Ensure the course belongs to the teacher's department
            if question.course.department == department:
                question.save()
        else:
            print("form is invalid")
        return HttpResponseRedirect('/teacher/teacher-view-question')
    return render(request, 'teacher/teacher_add_question.html', {'questionForm': questionForm})

@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def teacher_view_question_view(request):
    teacher = models.Teacher.objects.get(user=request.user)
    department = teacher.department
    courses = QMODEL.Course.objects.filter(department=department)
    return render(request, 'teacher/teacher_view_question.html', {'courses': courses})

@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def see_question_view(request, pk):
    teacher = models.Teacher.objects.get(user=request.user)
    department = teacher.department
    # Only allow viewing questions for courses in teacher's department
    try:
        course = QMODEL.Course.objects.get(id=pk, department=department)
        questions = QMODEL.Question.objects.filter(course=course)
    except QMODEL.Course.DoesNotExist:
        questions = QMODEL.Question.objects.none()
    return render(request, 'teacher/see_question.html', {'questions': questions})

@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def remove_question_view(request, pk):
    teacher = models.Teacher.objects.get(user=request.user)
    department = teacher.department
    try:
        question = QMODEL.Question.objects.get(id=pk)
        if question.course.department == department:
            question.delete()
    except QMODEL.Question.DoesNotExist:
        pass
    return HttpResponseRedirect('/teacher/teacher-view-question')

def logout(request):
    from django.contrib.auth import logout
    from django.contrib import messages
    logout(request)
    messages.info(request, 'logout successfully')
    return redirect('/')

@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def teacher_bulk_upload_questions_view(request):
    teacher = models.Teacher.objects.get(user=request.user)
    department = teacher.department
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        course_id = request.POST.get('course_id')
        try:
            course = QMODEL.Course.objects.get(id=course_id, department=department)
        except QMODEL.Course.DoesNotExist:
            return render(request, 'teacher/bulk_upload.html', {
                'courses': QMODEL.Course.objects.filter(department=department),
                'error': 'Selected course not found or not in your department.'
            })

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            created = 0
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                question_text, option1, option2, option3, option4, answer, marks = row or (None,) * 7
                if not question_text:
                    continue
                if answer not in ['Option1', 'Option2', 'Option3', 'Option4']:
                    continue
                try:
                    marks_int = int(marks) if marks is not None else 1
                except Exception:
                    marks_int = 1
                QMODEL.Question.objects.create(
                    course=course,
                    question=str(question_text)[:600],
                    option1=str(option1 or '')[:200],
                    option2=str(option2 or '')[:200],
                    option3=str(option3 or '')[:200],
                    option4=str(option4 or '')[:200],
                    answer=answer,
                    marks=marks_int
                )
                created += 1

            course.question_number = QMODEL.Question.objects.filter(course=course).count()
            course.total_marks = QMODEL.Question.objects.filter(course=course).aggregate(Sum('marks'))['marks__sum'] or 0
            course.save()

            return render(request, 'teacher/bulk_upload.html', {
                'courses': QMODEL.Course.objects.filter(department=department),
                'success': f'Successfully imported {created} questions into {course.course_name}.'
            })
        except Exception:
            return render(request, 'teacher/bulk_upload.html', {
                'courses': QMODEL.Course.objects.filter(department=department),
                'error': 'Failed to process file. Ensure it follows the sample format.'
            })

    return render(request, 'teacher/bulk_upload.html', {
        'courses': QMODEL.Course.objects.filter(department=department)
    })

@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def teacher_download_sample_questions_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Questions'
    ws.append(['question', 'option1', 'option2', 'option3', 'option4', 'answer', 'marks'])
    ws.append(['What is 2 + 2?', '1', '2', '3', '4', 'Option4', 2])
    ws.append(['Capital of France?', 'Berlin', 'Madrid', 'Paris', 'Lisbon', 'Option3', 5])
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    response = HttpResponse(stream.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sample_questions.xlsx"'
    return response